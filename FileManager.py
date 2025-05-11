import os
import shutil
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sns

class FileManager:
    def __init__(self, param_dict):

        self.param_dict = param_dict
        self.name = self.param_dict.get('name')

        # Prepare output directory and files
        self.output_dir = self.create_output_directory()
        self.output_file = os.path.join(self.output_dir, 'output.txt')
        self.excel_file = os.path.join(self.output_dir, 'data_output.xlsx')

        # Initialize necessary files and copy source files
        self.initialize_output_file()
        self.copy_source_files(os.getcwd())

    def create_output_directory(self):
        now = datetime.now().strftime('%Y-%m-%d_%H-%M')
        output_dir = os.path.join('Results', f'{now}_{self.name}')
        os.makedirs(output_dir, exist_ok=True)
        return output_dir

    def initialize_output_file(self):
        """Clears or initializes the output.txt file."""
        with open(self.output_file, 'w') as f:
            f.write("")

    def copy_source_files(self, source_dir):
        """
        Copies all files from the source directory to the output directory.
        Raises a FileNotFoundError if the source directory does not exist.
        """
        if not os.path.exists(source_dir):
            raise FileNotFoundError(f"The source directory {source_dir} does not exist.")

        for file_name in os.listdir(source_dir):
            full_file_name = os.path.join(source_dir, file_name)
            if os.path.isfile(full_file_name):
                shutil.copy(full_file_name, self.output_dir)

    def log(self, message):
        with open(self.output_file, 'a') as f:
            f.write(f"{message}\n")
        print(message)

    def log_model_structure(self, model):
        """Logs the model structure."""
        self.log(f"Model Structure:\n{model}")

    def log_tensor_dimensions(self, tensor_dict):
        """Logs the dimensions of tensors in the provided dictionary."""
        for tensor_name, tensor in tensor_dict.items():
            self.log(f"{tensor_name} dimensions: {list(tensor.size())}")

    def log_total_time(self, total_time):
        """Logs the total training time."""
        self.log(f"Total training time: {total_time:.2f} seconds")

    def log_sif_result(self, SIF_exact, SIF, SIF_error):
        """Logs the SIF (Stress Intensity Factor) results."""
        self.log(f"SIF_exact = {SIF_exact:.4f}    SIF_num = {SIF:.4f}    SIF_error = {SIF_error:.4E}")

    def print_and_save_losses(self, epoch, epoch_time, pde_loss_val, dir_loss_val, neumann_loss_val, interface_loss_val,
                              total_loss_val):

        # Round the values to 4 decimal places
        pde_loss_val = np.round(pde_loss_val.cpu().detach().numpy(), 10)
        dir_loss_val = np.round(dir_loss_val.cpu().detach().numpy(), 10)
        neumann_loss_val = np.round(neumann_loss_val.cpu().detach().numpy(), 10)
        interface_loss_val = np.round(interface_loss_val.cpu().detach().numpy(), 10)
        total_loss_val = np.round(total_loss_val.cpu().detach().numpy(), 10)

        # Print the losses
        output = (
            f'\n...Epoch {epoch + 1}...\n'
            f"PDE Loss:.........{pde_loss_val:.10f}\n"
            f"Dirichlet Loss:...{dir_loss_val:.10f}\n"
            f"Neumann Loss:.....{neumann_loss_val:.10f}\n"
            f"Interface Loss:...{interface_loss_val:.10f}\n"
            f"Total Loss:.......{total_loss_val:.10f}\n"
        )
        print(output)

        # Prepare the data for Excel
        data = {
            'Epoch': [epoch + 1],
            'PDE Loss': [pde_loss_val],
            'Dirichlet Loss': [dir_loss_val],
            'Neumann Loss': [neumann_loss_val],
            'Interface Loss': [interface_loss_val],
            'Total Loss': [total_loss_val],
            'epoch_duration': [epoch_time]
        }

        # Save the data to Excel
        self.save_to_excel(data, 'Losses')

    def write_to_excel(self, epoch, dlx_up, dly_up, dlx_down, dly_down, u1_up, u2_up, u1_down, u2_down):
        # Ensure the Excel file's directory exists
        os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)

        # Create a dataframe for each tensor with epoch
        data = {
            'epoch': [epoch + 1] * len(dlx_up),
            "dlx_up": dlx_up.cpu().detach().numpy().flatten(),
            "dly_up": dly_up.cpu().detach().numpy().flatten(),
            "dlx_down": dlx_down.cpu().detach().numpy().flatten(),
            "dly_down": dly_down.cpu().detach().numpy().flatten(),
            "u1_up": u1_up.cpu().detach().numpy().flatten(),
            "u2_up": u2_up.cpu().detach().numpy().flatten(),
            "u1_down": u1_down.cpu().detach().numpy().flatten(),
            "u2_down": u2_down.cpu().detach().numpy().flatten(),
        }

        # Save the data to Excel
        self.save_to_excel(data, 'input_output')

    def enrich_param_to_excel(self, epoch, enrich_param_up, enrich_param_down, enrich_param_up_grad,
                              enrich_param_down_grad):

        # Prepare the data for Excel
        data = {
            'Epoch': [epoch + 1],
            "enrich_param_up": enrich_param_up.cpu().detach().numpy().flatten(),
            "enrich_param_down": enrich_param_down.cpu().detach().numpy().flatten(),
            "enrich_param_up_grad": enrich_param_up_grad.cpu().detach().numpy().flatten(),
            "enrich_param_down_grad": enrich_param_down_grad.cpu().detach().numpy().flatten(),

        }

        # Save the data to Excel
        self.save_to_excel(data, 'enrich_param')

    def extend_function_excel(self, u1_add_up, u2_add_up, u1_x_add_up, u1_y_add_up, u2_x_add_up, u2_y_add_up,
                              u1_add_down, u2_add_down, u1_x_add_down, u1_y_add_down, u2_x_add_down, u2_y_add_down):
        # Store all variables in a dictionary
        data = {
            "u1_add_up": u1_add_up.cpu().detach().numpy().flatten(),
            "u2_add_up": u2_add_up.cpu().detach().numpy().flatten(),
            "u1_x_add_up": u1_x_add_up.cpu().detach().numpy().flatten(),
            "u1_y_add_up": u1_y_add_up.cpu().detach().numpy().flatten(),
            "u2_x_add_up": u2_x_add_up.cpu().detach().numpy().flatten(),
            "u2_y_add_up": u2_y_add_up.cpu().detach().numpy().flatten(),

            "u1_add_down": u1_add_down.cpu().detach().numpy().flatten(),
            "u2_add_down": u2_add_down.cpu().detach().numpy().flatten(),
            "u1_x_add_down": u1_x_add_down.cpu().detach().numpy().flatten(),
            "u1_y_add_down": u1_y_add_down.cpu().detach().numpy().flatten(),
            "u2_x_add_down": u2_x_add_down.cpu().detach().numpy().flatten(),
            "u2_y_add_down": u2_y_add_down.cpu().detach().numpy().flatten(),
        }

        # Save the data to Excel
        self.save_to_excel(data, 'extend_function')

    def save_to_excel(self, data, sheet_name):
        df = pd.DataFrame(data)
        # Check if the Excel file exists
        if os.path.exists(self.excel_file):
            with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Load the existing workbook
                book = load_workbook(self.excel_file)

                # Check if the 'Losses' sheet exists
                if sheet_name in book.sheetnames:
                    # If it exists, append the new data
                    startrow = book[sheet_name].max_row  # Start at the end of the existing data
                    df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow, header=False)
                else:
                    # If it doesn't exist, create a new sheet
                    df.to_excel(writer, index=False, sheet_name=sheet_name, header=True)
        else:
            # If the file doesn't exist, create it and write the data
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name, header=True)

    def plot_geometric_data(self, bp1_up, bp2_up, bp3_up, bp4_up, inp_up,
                            bp1_down, bp2_down, bp3_down, bp4_down, inp_down):

        # Get 10 distinct colors from the 'tab10' colormap
        colors = plt.cm.get_cmap('tab10').colors

        # Define data dictionary with different markers
        data_dict = {
            "bp1_up": (bp1_up, "o", colors[0]),
            "bp2_up": (bp2_up, "o", colors[1]),
            "bp3_up": (bp3_up, "o", colors[2]),
            "bp4_up": (bp4_up, "o", colors[3]),
            "inp_up": (inp_up, ".", colors[4]),
            "bp1_down": (bp1_down, "^", colors[5]),
            "bp2_down": (bp2_down, "^", colors[6]),
            "bp3_down": (bp3_down, "^", colors[7]),
            "bp4_down": (bp4_down, "^", colors[8]),
            "inp_down": (inp_down, ".", colors[4])
        }

        # Plot each set of points with a different marker
        plt.figure(figsize=(10,8))
        for label, (data, marker, color) in data_dict.items():
            plt.scatter(data[:, 0], data[:, 1], label=label, marker=marker, color=color)

        # Formatting the plot
        plt.xlabel("X Coordinate")
        plt.ylabel("Y Coordinate")
        plt.title("Scatter Plot of Input Data Point")
        plt.legend(loc="upper left", bbox_to_anchor=(1.01, 1), borderaxespad=0)
        plt.grid(True)
        # Adjust layout to prevent overlap
        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'input.png'))

        plt.show()

    def plot_results(self, dlx_up, dly_up, dlx_down, dly_down, u1_up, u2_up, u1_down, u2_down):
        # Convert tensors to numpy arrays
        dlx_up = dlx_up.detach().cpu().numpy()
        dly_up = dly_up.detach().cpu().numpy()
        dlx_down = dlx_down.detach().cpu().numpy()
        dly_down = dly_down.detach().cpu().numpy()
        u1_up = u1_up.detach().cpu().numpy()
        u2_up = u2_up.detach().cpu().numpy()
        u1_down = u1_down.detach().cpu().numpy()
        u2_down = u2_down.detach().cpu().numpy()

        # Compute displacement magnitude: U_mag = sqrt(Ux^2 + Uy^2)
        U_mag_up = np.sqrt(u1_up ** 2 + u2_up ** 2)
        U_mag_down = np.sqrt(u1_down ** 2 + u2_down ** 2)

        # Create DataFrames
        # Round x and y for cleaner pivot table labels
        df_up = pd.DataFrame({
            'x': np.round(dlx_up.flatten(), 2),
            'y': np.round(dly_up.flatten(), 2),
            'U_mag': U_mag_up.flatten()
        })

        df_down = pd.DataFrame({
            'x': np.round(dlx_down.flatten(), 2),
            'y': np.round(dly_down.flatten(), 2),
            'U_mag': U_mag_down.flatten()
        })
        # Pivot for heatmaps
        U_mag_pivot_up = df_up.pivot(index='y', columns='x', values='U_mag')
        U_mag_pivot_down = df_down.pivot(index='y', columns='x', values='U_mag')

        # Plot heatmaps
        fig, axes = plt.subplots(1, 2, figsize=(12, 6))

        sns.heatmap(U_mag_pivot_up, ax=axes[0], cmap="coolwarm", cbar=True)
        axes[0].set_title('Displacement Magnitude (Upper)')
        axes[0].invert_yaxis()  # Flip Y-axis

        sns.heatmap(U_mag_pivot_down, ax=axes[1], cmap="coolwarm", cbar=True)
        axes[1].set_title('Displacement Magnitude (Lower)')
        axes[1].invert_yaxis()  # Flip Y-axis

        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'combined_displacement_heatmap.png'))
        plt.show()

    def plot_losses(self, loss_vals):
        # Create a figure for Total Loss
        fig_total, ax_total = plt.subplots(figsize=(7, 5))  # Separate plot for Total Loss
        ax_total.plot(loss_vals['total'], '-', label='Total Loss')
        ax_total.set_yscale('log')  # Set y-axis to log scale
        ax_total.set_title('Total Loss')
        ax_total.set_xlabel('Epochs')
        ax_total.set_ylabel('Loss')
        ax_total.grid()
        ax_total.legend()

        # Save the Total Loss plot
        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'total_loss.png'))
        plt.show()

        # Create a figure for the other losses (PDE, Dirichlet, Neumann, Interface)
        fig, axs = plt.subplots(2, 2, figsize=(12, 8))  # 2x2 grid of subplots for the remaining losses
        axs = axs.ravel()  # Flatten the 2D array of axes to 1D for easy iteration

        # List of remaining loss types and their corresponding labels
        loss_types = ['pde', 'dir', 'neumann', 'interface']
        labels = ['PDE Loss', 'Dirichlet Loss', 'Neumann Loss', 'Interface Loss']

        for i, loss_type in enumerate(loss_types):
            axs[i].plot(loss_vals[loss_type], '-', label=labels[i])
            axs[i].set_yscale('log')  # Set y-axis to log scale for each subplot
            axs[i].set_title(f'{labels[i]}')
            axs[i].set_xlabel('Epochs')
            axs[i].set_ylabel('Loss')
            axs[i].grid()
            axs[i].legend()

        # Save the remaining losses plot
        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'other_losses.png'))
        plt.show()

    def plot_sif(self, r_kk, SIF_num, x_i, SIF_1_new):
        plt.plot(r_kk, SIF_num, 'b-o', label = 'SIF (numerical)')
        plt.plot(x_i, SIF_1_new, 'r--', label = 'SIF (extrapolated)')
        plt.ylabel('SIF')
        plt.xlabel('Distance from crack tip (r/a)')
        plt.title("Stress Intensity Factor")
        plt.legend()
        plt.savefig(os.path.join(self.output_dir, 'SIF_result.png'))
        plt.show()
